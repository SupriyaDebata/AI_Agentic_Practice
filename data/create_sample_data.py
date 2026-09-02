"""data/create_sample_data.py - Generate sample Excel and PDF for demo."""

from pathlib import Path


def create_excel() -> Path:
    """Create sample Excel workbook with 3 sheets."""
    from openpyxl import Workbook

    output = Path(__file__).parent / "sample_data.xlsx"
    wb = Workbook()

    # Sheet 1 - Quarterly Sales
    ws1 = wb.active
    ws1.title = "Quarterly Sales"
    ws1.append(["Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Annual Total"])
    
    rows = [
        ["Widget Alpha", 12000, 15000, 18000, 22000, 67000],
        ["Widget Beta", 8500, 11000, 13500, 16000, 49000],
        ["Gadget Pro", 25000, 28000, 31000, 35000, 119000],
        ["Gadget Lite", 5000, 7000, 9000, 11000, 32000],
        ["Service Pack", 3000, 3500, 4000, 4500, 15000],
        ["TOTAL", 53500, 64500, 75500, 88500, 282000],
    ]
    for row in rows:
        ws1.append(row)

    # Sheet 2 - Employee Data
    ws2 = wb.create_sheet("Employee Data")
    ws2.append(["ID", "Name", "Department", "Role", "Salary", "Start Date"])
    
    emp_rows = [
        [101, "Alice Johnson", "Engineering", "Senior Engineer", 95000, "2020-03-15"],
        [102, "Bob Smith", "Marketing", "Marketing Lead", 72000, "2019-07-01"],
        [103, "Carol White", "Engineering", "Data Scientist", 88000, "2021-01-10"],
        [104, "David Brown", "Sales", "Sales Director", 110000, "2018-05-20"],
        [105, "Eva Martinez", "HR", "HR Manager", 68000, "2022-02-28"],
        [106, "Frank Lee", "Engineering", "DevOps Engineer", 90000, "2021-09-01"],
        [107, "Grace Kim", "Finance", "CFO", 135000, "2017-11-15"],
    ]
    for row in emp_rows:
        ws2.append(row)

    # Sheet 3 - Inventory
    ws3 = wb.create_sheet("Inventory")
    ws3.append(["SKU", "Product Name", "Category", "Stock", "Reorder Level", "Unit Cost", "Status"])
    
    inv_rows = [
        ["WA-001", "Widget Alpha", "Widgets", 450, 100, 12.50, "OK"],
        ["WB-002", "Widget Beta", "Widgets", 280, 80, 9.75, "OK"],
        ["GP-010", "Gadget Pro", "Gadgets", 95, 25, 45.00, "LOW"],
        ["GL-011", "Gadget Lite", "Gadgets", 320, 75, 18.00, "OK"],
        ["SP-020", "Service Pack A", "Services", 60, 20, 35.00, "OK"],
        ["AC-030", "Accessory Kit", "Accessories", 800, 200, 5.50, "OK"],
    ]
    for row in inv_rows:
        ws3.append(row)

    wb.save(str(output))
    print(f"Created: {output}")
    return output


def create_pdf() -> Path:
    from fpdf import FPDF

    output = Path(__file__).parent / "sample_report.pdf"

    class PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 11)
            self.set_text_color(100, 100, 200)
            self.cell(0, 10, "TechCorp Annual Report 2024", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(2)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # -- Page 1: Executive Summary ---------------------------------------------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(30, 30, 80)
    pdf.cell(0, 12, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(50, 50, 50)
    for para in [
        (
            "TechCorp achieved record annual revenue of $282,000 in fiscal year 2024, "
            "representing a 23% increase over the previous year. This growth was driven "
            "primarily by strong performance in our Gadget Pro product line, which "
            "contributed $119,000 - over 42% of total revenue."
        ),
        (
            "Our Q3 2024 results were particularly strong, with total sales reaching "
            "$75,500 across all product categories. The Widget Alpha product line "
            "showed consistent quarter-over-quarter growth, rising from $12,000 in Q1 "
            "to $22,000 in Q4 2024."
        ),
        (
            "The company employs 7 full-time staff across Engineering, Marketing, Sales, "
            "HR, and Finance departments. Our CFO Grace Kim joined TechCorp in 2017 and "
            "has overseen the company's financial strategy through three consecutive "
            "years of revenue growth. Total payroll stands at approximately $658,000 annually."
        ),
        (
            "Looking ahead to 2025, TechCorp plans to expand the Gadget Pro line with two "
            "new variants and invest in automation to reduce fulfillment costs. "
            "The Inventory team has flagged that Gadget Pro stock (currently 95 units) "
            "is approaching its reorder level of 25 units."
        ),
    ]:
        pdf.multi_cell(0, 7, para)
        pdf.ln(4)

    # -- Page 2: Data Tables ---------------------------------------------------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 30, 80)
    pdf.cell(0, 10, "Financial Data Tables", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, "Table 1 - Quarterly Sales by Product (USD)", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    col_widths = [45, 25, 25, 25, 25, 30]
    headers    = ["Product", "Q1", "Q2", "Q3", "Q4", "Total"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(80, 80, 180)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    table_data = [
        ("Widget Alpha",  "12,000", "15,000", "18,000", "22,000",  "67,000"),
        ("Widget Beta",    "8,500", "11,000", "13,500", "16,000",  "49,000"),
        ("Gadget Pro",    "25,000", "28,000", "31,000", "35,000", "119,000"),
        ("Gadget Lite",    "5,000",  "7,000",  "9,000", "11,000",  "32,000"),
        ("Service Pack",   "3,000",  "3,500",  "4,000",  "4,500",  "15,000"),
        ("TOTAL",         "53,500", "64,500", "75,500", "88,500", "282,000"),
    ]
    pdf.set_font("Helvetica", "", 9)
    for i, row in enumerate(table_data):
        is_total = i == len(table_data) - 1
        if is_total:
            pdf.set_fill_color(220, 220, 240)
            pdf.set_font("Helvetica", "B", 9)
        else:
            pdf.set_fill_color(245 if i % 2 == 0 else 255, 245 if i % 2 == 0 else 255, 252 if i % 2 == 0 else 255)
            pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(30, 30, 30)
        for w, val in zip(col_widths, row):
            pdf.cell(w, 7, val, border=1, fill=True)
        pdf.ln()

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, "Table 2 - Inventory Status", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    inv_hdrs = ["SKU", "Product Name", "Stock", "Reorder Lvl", "Status"]
    inv_wids = [22, 45, 22, 30, 36]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(80, 180, 120)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(inv_wids, inv_hdrs):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    inv_data = [
        ("WA-001", "Widget Alpha",   "450", "100", "OK"),
        ("WB-002", "Widget Beta",    "280",  "80", "OK"),
        ("GP-010", "Gadget Pro",      "95",  "25", "LOW"),
        ("GL-011", "Gadget Lite",    "320",  "75", "OK"),
        ("SP-020", "Service Pack A",  "60",  "20", "OK"),
        ("AC-030", "Accessory Kit",  "800", "200", "OK"),
    ]
    pdf.set_font("Helvetica", "", 9)
    for i, row in enumerate(inv_data):
        pdf.set_fill_color(245 if i % 2 == 0 else 255, 252 if i % 2 == 0 else 255, 245 if i % 2 == 0 else 255)
        pdf.set_text_color(30, 30, 30)
        for w, val in zip(inv_wids, row):
            pdf.cell(w, 7, val, border=1, fill=True)
        pdf.ln()

    # -- Page 3: Strategic Outlook ---------------------------------------------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 30, 80)
    pdf.cell(0, 10, "Strategic Outlook 2025", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, "Key Initiatives", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(50, 50, 50)
    for item in [
        "1. Launch Gadget Pro XL and Gadget Pro Mini by Q2 2025",
        "2. Automate warehouse operations to cut fulfillment costs by 15%",
        "3. Expand the Service Pack portfolio to five tiers",
        "4. Hire three additional engineers to support new product development",
        "5. Open a second distribution centre in the East region",
    ]:
        pdf.cell(0, 8, item, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, "Risk Factors", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(50, 50, 50)
    pdf.multi_cell(0, 7,
        "Supply chain disruptions remain a key risk for the Gadget Pro line. "
        "The current stock of 95 units provides approximately 3 weeks of buffer "
        "at current demand rates. The procurement team has been authorised to "
        "increase safety stock to 150 units by January 2025.\n\n"
        "Foreign exchange exposure affects approximately 30% of our raw material "
        "costs. The Finance team, led by CFO Grace Kim, has implemented a hedging "
        "strategy covering the next two quarters."
    )
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, "Revenue Target 2025", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(50, 50, 50)
    pdf.multi_cell(0, 7,
        "The Board has approved a revenue target of $360,000 for fiscal year 2025, "
        "representing a 27.7% increase over 2024 actuals. "
        "This target assumes successful launch of both new Gadget Pro variants "
        "and a full-year contribution from the East distribution centre."
    )

    pdf.output(str(output))
    print(f"Created: {output}")
    return output


if __name__ == "__main__":
    data_dir = Path(__file__).parent
    data_dir.mkdir(exist_ok=True)

    print("Generating sample data files...")
    xlsx_path = create_excel()
    pdf_path  = create_pdf()

    print()
    print("=" * 58)
    print("Sample files ready:")
    print(f"  Excel : {xlsx_path}")
    print(f"  PDF   : {pdf_path}")
    print()
    print("Next steps:")
    print("  1. Start Ollama :  ollama serve")
    print("  2. Run the app  :  streamlit run app.py")
    print("  3. Upload both files from the data/ folder")
    print("  4. Try asking:")
    print('       "What is the total annual revenue?"')
    print('       "Who is the CFO?"')
    print('       "Which product has the lowest stock?"')
    print('       "What is Alice\'s salary?"')
    print('       "Does the document have any charts?"')
    print("=" * 58)
